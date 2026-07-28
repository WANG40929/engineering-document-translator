from __future__ import annotations

import time

from openpyxl import load_workbook

from ..i18n import tr
from ..models import FileResult
from ..text_utils import is_translatable
from .base import TranslationEngine


class XlsxEngine(TranslationEngine):
    extensions = (".xlsx", ".xlsm")

    def translate(self, source, destination, translator, options, progress=None) -> FileResult:
        started = time.monotonic()
        result = FileResult(str(source), str(destination), engine="XLSX")
        try:
            keep_vba = source.suffix.lower() == ".xlsm"
            workbook = load_workbook(source, keep_vba=keep_vba, data_only=False)
            cells = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and not cell.value.startswith("=") and is_translatable(cell.value):
                            cells.append(cell)
            texts = [cell.value for cell in cells]

            def report(done, total):
                if progress:
                    progress(
                        str(source),
                        done / max(total, 1),
                        tr("progress.excel", done=done, total=total),
                    )

            translations = translator.translate_many(texts, report)
            for cell, value in zip(cells, translations):
                cell.value = value
            destination.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(destination)
            workbook.close()
            result.translated_units = len(cells)
            result.status = "completed"
        except Exception as exc:
            result.status = "failed"
            result.errors.append(str(exc))
            if destination.exists():
                destination.unlink()
        result.elapsed_seconds = round(time.monotonic() - started, 2)
        result.usage = dict(getattr(translator, "usage", {}))
        return result

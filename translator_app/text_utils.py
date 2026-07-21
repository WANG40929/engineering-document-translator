from __future__ import annotations

import csv
import hashlib
import re
from dataclasses import dataclass
from pathlib import Path


LETTER_RE = re.compile(r"[A-Za-z\u00c0-\u024f\u0400-\u04ff\u4e00-\u9fff\u3040-\u30ff\uac00-\ud7af]")
ONLY_CODE_RE = re.compile(r"^[\s\d\W_]+$", re.UNICODE)
PROTECTED_PATTERNS = [
    # URLs, e-mail addresses and explicit placeholders.
    r"https?://\S+",
    r"\b[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}\b",
    r"__UDT_\d{4}__",
    # Standards, drawing/KKS identifiers and long mixed codes.
    r"\b(?:ISO|IEC|DIN|EN|GB|ASTM|ASME|API|IEEE|NFPA|GOST|SNIP|SP|SNiP)[ -]?[A-Z0-9][A-Z0-9./:_-]*\b",
    r"\b(?=[A-Z0-9&._/-]{5,}\b)(?=[A-Z0-9&._/-]*\d)(?=[A-Z0-9&._/-]*[A-Z])[A-Z0-9&._/-]+\b",
    # Dimensions, temperatures, pressures, electrical values and percentages.
    r"(?<!\w)[+-]?\d+(?:[.,]\d+)?\s*(?:mm|cm|m|km|kg|g|t|N|kN|Nm|kNm|Pa|kPa|MPa|bar|psi|°C|K|V|kV|A|mA|Hz|kW|MW|rpm|r/min|%)(?!\w)",
    r"(?<!\w)\d+(?:[.,]\d+)?\s*[x×]\s*\d+(?:[.,]\d+)?(?:\s*[x×]\s*\d+(?:[.,]\d+)?)?(?!\w)",
]
PROTECTED_RE = re.compile("|".join(f"(?:{p})" for p in PROTECTED_PATTERNS), re.IGNORECASE)
PLACEHOLDER_FLEX_RE = re.compile(
    r"(?<![A-Za-z0-9])(?:_{1,2}\s*)?UDT\s*_?\s*(\d{4})(?:\s*_{1,2})?(?![A-Za-z0-9])",
    re.IGNORECASE,
)


def normalize_text(text: str) -> str:
    return re.sub(r"[ \t\u00a0]+", " ", text.replace("\r", "")).strip()


def is_translatable(text: str) -> bool:
    value = normalize_text(text)
    if not value or ONLY_CODE_RE.match(value) or not LETTER_RE.search(value):
        return False
    # If protection removes every word, the value is only a code/standard/unit.
    if not LETTER_RE.search(PROTECTED_RE.sub("", value)):
        return False
    # Single Latin letters are normally drawing references, not prose.
    if len(value) == 1 and value.isascii():
        return False
    return True


@dataclass(slots=True)
class ProtectedText:
    text: str
    values: list[str]

    def restore(self, translated: str) -> str:
        result = translated
        for index, value in enumerate(self.values):
            token = f"__UDT_{index:04d}__"
            # Models occasionally remove underscores or insert spaces inside
            # placeholders. Accept those variants so internal markers never
            # leak into a translated document.
            flexible = re.compile(
                rf"(?<![A-Za-z0-9])(?:_{{1,2}}\s*)?UDT\s*_?\s*{index:04d}(?:\s*_{{1,2}})?(?![A-Za-z0-9])",
                re.IGNORECASE,
            )
            result = flexible.sub(lambda _m, v=value: v, result)
            result = result.replace(token, value)
        return result


def placeholder_indexes(text: str) -> list[int]:
    """Return all internal placeholder indexes, including model-mangled forms."""
    return [int(match.group(1)) for match in PLACEHOLDER_FLEX_RE.finditer(text)]


def has_internal_placeholder(text: str) -> bool:
    return PLACEHOLDER_FLEX_RE.search(text) is not None


def protect_text(text: str) -> ProtectedText:
    values: list[str] = []

    def replace(match: re.Match) -> str:
        values.append(match.group(0))
        return f"__UDT_{len(values) - 1:04d}__"

    return ProtectedText(PROTECTED_RE.sub(replace, text), values)


def load_glossary(path: Path | None) -> dict[str, str]:
    if not path:
        return {}
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"术语表不存在：{path}")
    result: dict[str, str] = {}
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        book = load_workbook(path, read_only=True, data_only=True)
        sheet = book.active
        for row in sheet.iter_rows(values_only=True):
            if len(row) >= 2 and row[0] and row[1]:
                result[str(row[0]).strip()] = str(row[1]).strip()
        book.close()
        return result
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;=")
        except csv.Error:
            dialect = csv.excel
        for row in csv.reader(handle, dialect):
            if len(row) >= 2 and row[0].strip() and row[1].strip():
                result[row[0].strip()] = row[1].strip()
    return result


def glossary_signature(glossary: dict[str, str]) -> str:
    joined = "\n".join(f"{k}\t{v}" for k, v in sorted(glossary.items()))
    return hashlib.sha256(joined.encode("utf-8")).hexdigest() if joined else ""
